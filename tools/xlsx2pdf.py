#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel (.xlsx) -> PDF: tüm sayfaları, satır/sütun/içerik, birleşik hücreler
(SPAN) ve Türkçe karakterler korunarak birebir basar. Sütun genişlikleri içeriğe
göre otomatik; sayısal sütunlar sağa yaslı; metin kıvrılır. LibreOffice gerekmez.

Kullanım: python3 tools/xlsx2pdf.py girdi.xlsx cikti.pdf
"""
import sys
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, LongTable, TableStyle, Paragraph, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

pdfmetrics.registerFont(TTFont("DV", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
pdfmetrics.registerFont(TTFont("DVB", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"))


def tr_num(v):
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int) or (isinstance(v, float) and float(v).is_integer()):
        return f"{int(v):,d}".replace(",", ".")
    s = f"{v:,.4f}"
    intp, frac = s.split(".")
    frac = frac.rstrip("0")
    intp = intp.replace(",", ".")
    return intp + ("," + frac if frac else "")


def esc(t):
    return t.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return tr_num(v)
    return str(v)


def full_span_rows(ws, ncol):
    """min_col=1 ve max_col=ncol olan birleşik satırların kümesi."""
    s = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 1 and mr.max_col == ncol:
            for r in range(mr.min_row, mr.max_row + 1):
                s[r] = mr.min_row
    return s


def col_weights(ws, ncol, fullspans):
    maxlen = [3] * ncol
    numeric_hits = [0] * ncol
    numeric_total = [0] * ncol
    for r in range(1, ws.max_row + 1):
        if r in fullspans:
            continue
        for c in range(1, ncol + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            txt = cell_str(v)
            # uzun açıklama sütununu biraz sınırla ama görünür tut
            maxlen[c - 1] = max(maxlen[c - 1], min(len(txt), 70))
            numeric_total[c - 1] += 1
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                numeric_hits[c - 1] += 1
    weights = [max(3.0, ml) for ml in maxlen]
    right = {i for i in range(ncol)
             if numeric_total[i] >= 3 and numeric_hits[i] >= 0.6 * numeric_total[i]}
    return weights, right


def build_sheet_story(ws):
    ncol = ws.max_column
    nrow = ws.max_row
    if ncol == 0 or nrow == 0:
        return []
    fullspans = full_span_rows(ws, ncol)
    weights, right_cols = col_weights(ws, ncol, fullspans)

    page_w = landscape(A4)[0] - 16 * mm
    tot = sum(weights)
    col_w = [page_w * w / tot for w in weights]

    base = ParagraphStyle("b", fontName="DV", fontSize=6.4, leading=7.8)
    st_left = ParagraphStyle("l", parent=base, alignment=0)
    st_right = ParagraphStyle("r", parent=base, alignment=2)
    st_banner = ParagraphStyle("ban", parent=base, alignment=0, fontName="DVB", fontSize=6.8)
    st_title = ParagraphStyle("t", parent=base, alignment=0, fontName="DVB",
                              fontSize=9, leading=11, textColor=colors.HexColor("#1f4e79"))

    data = []
    styles = []
    spans = []
    for r in range(1, nrow + 1):
        is_full = r in fullspans
        is_title = (r == 1)
        row = []
        for c in range(1, ncol + 1):
            v = ws.cell(r, c).value
            txt = esc(cell_str(v))
            if is_title and c == 1:
                stl = st_title
            elif is_full and c == 1:
                stl = st_banner
            elif (c - 1) in right_cols and not is_full:
                stl = st_right
            else:
                stl = st_left
            row.append(Paragraph(txt or "&nbsp;", stl))
        data.append(row)
        if is_full:
            spans.append(("SPAN", (0, r - 1), (ncol - 1, r - 1)))
            styles.append(("BACKGROUND", (0, r - 1), (ncol - 1, r - 1),
                           colors.HexColor("#eaf0f7" if r > 1 else "#ffffff")))

    ts = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c2c7d0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
    ] + spans + styles)

    tbl = LongTable(data, colWidths=col_w)
    tbl.setStyle(ts)
    return [tbl]


def main(src, out):
    wb = load_workbook(src, data_only=True)
    story = []
    for i, ws in enumerate(wb.worksheets):
        if i > 0:
            story.append(PageBreak())
        story.extend(build_sheet_story(ws))
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=8 * mm, rightMargin=8 * mm,
                            topMargin=8 * mm, bottomMargin=8 * mm,
                            title="Belge")
    doc.build(story)
    print("PDF üretildi:", out)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
